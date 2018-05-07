from backports import ssl
from imapclient import IMAPClient
import pyzmail
import imaplib
import bs4
import getpass
import calendar
import openpyxl
import datetime
from re import sub
from decimal import Decimal


# ---------------------------------------------------------------------------------
# Definitions
# ---------------------------------------------------------------------------------
imaplib._MAXLINE = 10000000

IMAP_SERVER  = 'imap.gmail.com'
USERNAME     = 'user@gmail.com'

# For the spreadsheet
COLUMN_MONTH  = 1
COLUMN_UBER   = 2
COLUMN_CABIFY = 3
COLUMN_99     = 4
COLUMN_TOTAL  = 5
ROW_OFFSET    = 3

# Months and year of search
FIRST_MONTH = 4
LAST_MONTH  = 5
YEAR = 2018

OUTPUT_FILE = "imap_get_rides.xlsx"

'''
Examples for searching:
after:2017/11/01 before:2017/12/01 trip with uber
after:2017/11/01 before:2017/12/01 journey with cabify
after:2017/11/01 before:2017/12/01 from:nao-responder@99taxis.com
'''

def getRidesCost(company, month):
    AFTER  = "{}/{}/31".format(YEAR, month-1)
    BEFORE = "{}/{}/01".format(YEAR, month+1)

    TO_SEARCH_UBER   = "after:%s before:%s 'trip with uber'" % (AFTER,BEFORE)
    TO_SEARCH_CABIFY = "after:%s before:%s 'journey with cabify'" % (AFTER,BEFORE)
    TO_SEARCH_99     = "after:%s before:%s 'sua corrida de hoje com a 99'" % (AFTER,BEFORE)

    UBER_TAG   = "td"
    UBER_CLASS = "totalPrice topPrice tal black"

    CABIFY_TAG   = "h1"
    CABIFY_CLASS = "heading-title txt-right"

    _99_TAG   = "h1"
    _99_CLASS = "m_8307221800254057134align-right m_8307221800254057134no-wrap"

    total = 0
    UIDs  = []
    print("Searching for rides with %s in %s" % (company, calendar.month_name[month]) )

    # Search
    if company == "Uber":
        UIDs = imapObj.gmail_search(TO_SEARCH_UBER)
        print(TO_SEARCH_UBER)
    elif company == "Cabify":
        UIDs = imapObj.gmail_search(TO_SEARCH_CABIFY)
        print(TO_SEARCH_CABIFY)
    else: #if company == "99":
        UIDs = imapObj.gmail_search(TO_SEARCH_99)
        print(TO_SEARCH_99)

    print("%d ride(s) found" % len(UIDs ))

    # Fetch
    rawMessages = imapObj.fetch(UIDs, ['BODY[]'])

    for id in UIDs:
        message = pyzmail.PyzMessage.factory(rawMessages[id][b'BODY[]'])

        try:

            if message.html_part != None:
                page = bs4.BeautifulSoup(message.html_part.get_payload().decode(message.html_part.charset), "html.parser") #"lxml"

                if company == "Uber":
                    value = page.find_all(UBER_TAG, class_=UBER_CLASS)
                elif company == "Cabify":
                    value = page.find_all(CABIFY_TAG, class_=CABIFY_CLASS)
                else:
                    table = page.find_all('table', class_=lambda x: x and 'payment' in x.split('-'))

                    # This is the cell we want
                    cell = table[0].find_next('tbody').find_all('tr')[-1].find_all('td')[1]

                    # We want the second cell
                    value = cell.text.replace('R$', '')

                if value == []:
                    print('Could not find value') #TODO NOT reliable??
                else:
                    if company == "99":
                        ride = Decimal(value.replace(',', '.'))
                    else:
                        ride = Decimal(sub(r'[^\d.]', '', value[0].contents[0]))

                    print(ride)
                    total += ride
        except:
            print("Something wrong")

    print("Rides with %s: %.2f\n" % (company, total))
    return total

""" Connect with IMAP """
# IMAP Connection - workaround (http://stackoverflow.com/a/40209809/660711)
try:
    context = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
    imapObj = IMAPClient(IMAP_SERVER, ssl=True, ssl_context=context)
except:
    print("Exception connecting to server '%s'" % IMAP_SERVER)
    quit()

# ---------------------------------------------------------------------------------
# Login
# ---------------------------------------------------------------------------------
_pswd  = getpass.getpass('Password: ')

try:
    imapObj.login(USERNAME, _pswd)
except:
    print("Exception logging in user '%s'. Wrong password?" % USERNAME)
    quit()

# Select a folder
imapObj.select_folder('INBOX', readonly=True)

# ---------------------------------------------------------------------------------
# Write the results to a excel spreadsheet
# ---------------------------------------------------------------------------------
# Create a new workbook
wb = openpyxl.Workbook()

# Grab the active worksheet
sheet = wb.active

# Change the name of the worksheet
sheet.title = 'Uber and Cabify rides'

# Basic information
sheet['A1'] = datetime.datetime.now()
sheet['B1'] = USERNAME
sheet['B3'] = "Uber"
sheet['C3'] = "Cabify"
sheet['D3'] = "Total"


# Save monthly values
by_month = []

for month in range(FIRST_MONTH, LAST_MONTH):

    sheet.cell(row=ROW_OFFSET+month, column=COLUMN_MONTH).value = calendar.month_name[month]

    """ Get all trips with uber in the month"""
    total_uber = getRidesCost("Uber", month)
    sheet.cell(row=ROW_OFFSET+month, column=COLUMN_UBER).value = total_uber

    """ Get all journeys with cabify in the month"""
    total_cabify = getRidesCost("Cabify", month)
    sheet.cell(row=ROW_OFFSET+month, column=COLUMN_CABIFY).value = total_cabify

    """ Get all journeys with 99 in the month"""
    total_99 = 0
    total_99 = getRidesCost("99", month)
    sheet.cell(row=ROW_OFFSET+month, column=COLUMN_99).value = total_99

    by_month.append(total_uber + total_cabify + total_99)
    sheet.cell(row=ROW_OFFSET+month, column=COLUMN_TOTAL).value = total_uber + total_cabify + total_99
    print("---> Total of rides in %s: %.2f\n" % (calendar.month_name[month], by_month[-1]) )


# ---------------------------------------------------------------------------------
# Print the result
# ---------------------------------------------------------------------------------
for i in range(FIRST_MONTH, LAST_MONTH):
    print("%s: %.2f" % (calendar.month_name[i], by_month[i-FIRST_MONTH]))

# ---------------------------------------------------------------------------------
# Save the file
# ---------------------------------------------------------------------------------
try:
    wb.save(OUTPUT_FILE)
except:
    print("Failed saving file. Does the file already exist?")

# ---------------------------------------------------------------------------------
# Logout
# ---------------------------------------------------------------------------------
imapObj.logout()
